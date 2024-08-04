import os
import pandas as pd
import openpyxl
import requests
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm

# Create the output directory if it doesn't exist
output_dir = 'output/link_accessibility'
os.makedirs(output_dir, exist_ok=True)

# Read the Excel file into a DataFrame
df = pd.read_excel('output/economist_audio_urls.xlsx')

# Create a new column to store the accessibility status
df['Accessibility'] = ''

# Define a function to check the accessibility of a URL
def check_url_accessibility(url):
    try:
        response = requests.head(url)
        if response.status_code == 200:
            return 'Accessible'
        else:
            return 'Not Accessible'
    except requests.exceptions.RequestException:
        return 'Error'

# Use ThreadPoolExecutor to run the URL accessibility checks concurrently
with ThreadPoolExecutor() as executor:
    # Create a list to store the future objects
    futures = []

    # Submit the URL checks as concurrent tasks
    for url in df['URL']:
        future = executor.submit(check_url_accessibility, url)
        futures.append(future)

    # Iterate through the futures to update the DataFrame
    with tqdm(total=len(futures)) as pbar:
        for index, future in enumerate(futures):
            accessibility = future.result()
            df.at[index, 'Accessibility'] = accessibility
            pbar.update(1)  # Update the progress bar

# Save the updated DataFrame to a new Excel file
xlsx_file_path = os.path.join(output_dir, 'economist_audio_urls_accessibility.xlsx')
df.to_excel(xlsx_file_path, index=False)

# Save the accessibility status as a text file with the same title
txt_title = os.path.join(output_dir, 'economist_audio_urls_accessibility.txt')
with open(txt_title, 'w') as txt_file:
    # Write the column titles to the text file
    txt_file.write("Date\tIssue\tURL\tAccessibility\n")

    # Write each row to the text file
    for _, row in df.iterrows():
        txt_file.write(f"{row['Date']}\t{row['Issue']}\t{row['URL']}\t{row['Accessibility']}\n")

# Adjust column widths in the resulting Excel file
workbook = openpyxl.load_workbook(xlsx_file_path)
worksheet = workbook.active

# Set column widths for all columns
for column_letter in worksheet.columns:
    max_length = 0
    column_index = column_letter[0].column  # Get the index of the column
    for cell in column_letter:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1  # Adjust the width slightly for better visibility
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = adjusted_width

# Add filters to all columns
last_column_index = worksheet.max_column
last_column_letter = openpyxl.utils.get_column_letter(last_column_index)
worksheet.auto_filter.ref = f"A1:{last_column_letter}1"

# Save the workbook with adjusted column widths
workbook.save(xlsx_file_path)
