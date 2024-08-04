import os
import pandas as pd
import openpyxl
from datetime import date

def is_first_saturday_in_august(date):
    return date.year >= 2022 and date.year != 2024 and date.month == 8 and date.weekday() == 5 and 1 <= date.day <= 7

year_input = input("Enter the year (or press Enter for the current year): ")

if year_input.strip() == "":
    year = date.today().year
else:
    year = int(year_input)

last_day = pd.Timestamp(year, 12, 31)

schedule_day = pd.date_range('20070526', last_day, freq='W-SAT')
weeks = 0
urls = []

for i in schedule_day:
    year = i.strftime('%Y')
    month = i.strftime('%m')
    day = i.strftime('%d')
    date_str = i.strftime('%Y%m%d')
    issue = 8796 - 266 + weeks
    if (month != '12' or day < '25') and not is_first_saturday_in_august(i):
        if date_str == '20190330':
            url = "https://audiocdn.economist.com/sites/default/files/AudioArchive/{0}/{2}/Issue_{1}_{2}_The_Economist_Full_Edition.zip".format(year, issue, date_str)
        elif date_str == '20210123':
            url = "https://audiocdn.economist.com/sites/default/files/AudioArchive/{0}/{2}/Issue_{1}_{2}_The_Economist_Full_edition.zip".format(year, issue+1, date_str)
        elif date_str == '20111224':
            new_date = str(int(date_str) + 7)
            url = "https://audiocdn.economist.com/sites/default/files/AudioArchive/{0}/{2}/{2}_TheEconomist_Full_Edition.zip".format(year, issue, new_date)
        elif date_str <= '20120728':
            url = "https://audiocdn.economist.com/sites/default/files/AudioArchive/{0}/{2}/{2}_TheEconomist_Full_Edition.zip".format(year, issue, date_str)
        else:
            url = "https://audiocdn.economist.com/sites/default/files/AudioArchive/{0}/{2}/Issue_{1}_{2}_The_Economist_Full_edition.zip".format(year, issue, date_str)
        urls.append((i, issue, url))
        weeks += 1

df = pd.DataFrame(urls, columns=['Date', 'Issue', 'URL'])
df['Date'] = pd.to_datetime(df['Date'])
df['Date'] = df['Date'].dt.strftime('%Y-%m-%d')

output_folder = 'output'
if not os.path.exists(output_folder):
    os.makedirs(output_folder)

excel_filename = os.path.join(output_folder, 'economist_audio_urls.xlsx')
df.to_excel(excel_filename, index=False)

txt_filename = os.path.join(output_folder, 'economist_audio_urls.txt')
with open(txt_filename, 'w') as txt_file:
    txt_file.write("Date\tIssue\tURL\n")
    for _, row in df.iterrows():
        txt_file.write(f"{row['Date']}\t{row['Issue']}\t{row['URL']}\n")

workbook = openpyxl.load_workbook(excel_filename)
worksheet = workbook.active

for column_letter in worksheet.columns:
    max_length = 0
    column_index = column_letter[0].column
    for cell in column_letter:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1
    worksheet.column_dimensions[openpyxl.utils.get_column_letter(column_index)].width = adjusted_width

last_column_index = worksheet.max_column
last_column_letter = openpyxl.utils.get_column_letter(last_column_index)
worksheet.auto_filter.ref = f"A1:{last_column_letter}1"

workbook.save(excel_filename)
