import requests
from bs4 import BeautifulSoup
from datetime import datetime
import pandas as pd
from tqdm import tqdm
import concurrent.futures
from dateutil import parser
import openpyxl
import re

base_url = 'https://www.economist.com/weeklyedition/archive?year='
start_year = 1997

current_year = datetime.now().year
end_year = current_year

data = []

# Function to fetch data (date and URL) for a given year
def fetch_data(year):
    url = base_url + str(year)
    response = requests.get(url)
    html_content = response.text

    soup = BeautifulSoup(html_content, 'html.parser')
    time_elements = soup.find_all('time', class_='edition-teaser__subheadline')
    url_elements = soup.find_all('meta', itemprop='url')

    year_data = []
    for time_element, url_element in zip(time_elements, url_elements):
        date = time_element.text.strip()
        date_obj = parser.parse(date, dayfirst=True)
        formatted_date = date_obj.strftime('%Y-%m-%d')
        url = url_element.get('content')
        year_data.append({'Date': formatted_date, 'URL': url})

    return year_data

# Display progress bar
def progress_bar(iterator):
    return tqdm(iterator, total=end_year - start_year + 1)

# Use multi-threading to fetch data
with concurrent.futures.ThreadPoolExecutor() as executor:
    results = list(progress_bar(executor.map(fetch_data, range(start_year, end_year + 1))))

# Flatten the results
data = [item for sublist in results for item in sublist]

# Create a DataFrame
df = pd.DataFrame(data)

# Replace the desired portion of the URLs using regular expressions, the largest resolution is width/height=11470/11470
df['URL'] = df['URL'].str.replace(r'https://www.economist.com/img/b/\d+/\d+/90/', 'https://www.economist.com/img/b/5000/5000/90/', regex=True)

# Sort the DataFrame in reverse alphabetical order of the 'Date' column
df = df.sort_values('Date', ascending=False)

# Export to Excel
writer = pd.ExcelWriter('IMG.xlsx', engine='xlsxwriter')
df.to_excel(writer, index=False)

# Close the Pandas Excel writer and get the workbook
writer.close()
workbook = openpyxl.load_workbook('IMG.xlsx')

# Get the worksheet
worksheet = workbook['Sheet1']

# Set column widths for all columns
columns = worksheet.columns
for column in columns:
    max_length = 0
    column_letter = column[0].column_letter
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except TypeError:
            pass
    adjusted_width = (max_length + 2) * 1  # Adjust the width slightly for better visibility
    worksheet.column_dimensions[column_letter].width = adjusted_width

# Add filters to all columns
last_column_index = worksheet.max_column
last_column_letter = openpyxl.utils.get_column_letter(last_column_index)
worksheet.auto_filter.ref = f"A1:{last_column_letter}1"

# Freeze the first row
worksheet.freeze_panes = 'A2'

# Save the workbook with adjusted column widths and filters
workbook.save('IMG.xlsx')

# Save data to a text file
with open('IMG.txt', 'w') as txt_file:
    for row_index in range(1, worksheet.max_row + 1):
        row_data = [str(worksheet.cell(row=row_index, column=col_index).value) for col_index in range(1, worksheet.max_column + 1)]
        txt_file.write('\t'.join(row_data) + '\n')

print("Data saved to Excel (IMG.xlsx) and text (IMG.txt) files.")